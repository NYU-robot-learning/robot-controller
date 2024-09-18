import pickle
import random
import io
from PIL import Image
import base64
import torch
import os
import time
import gc
import traceback
import openpyxl
from flask import Flask, render_template, jsonify, request
import pandas as pd

app = Flask(__name__)

page_size = 24
MAX_IMAGES = 96000
MAX_QUERIES = 10

class Environment:
    def __init__(self, env_number, time_steps, pkl_file, max_images=MAX_IMAGES, max_queries=MAX_QUERIES):
        self.env_number = env_number
        self.time_steps = time_steps
        self.pkl_file = pkl_file
        self.max_images = max_images
        self.max_queries = max_queries
        self.current_time_step = None
        self.env_data = {}
        self.queries = []
        self.completed_queries = []
        self.query_index = 0
        self.current_images = []
        self.previous_time_step = None

    def load_environment_images(self):
        folder_path = f'env{self.env_number}'
        pkl_file_path = os.path.join(folder_path, self.pkl_file)

        if not os.path.exists(pkl_file_path):
            raise FileNotFoundError(f"PKL file for environment {self.env_number} not found.")

        with open(pkl_file_path, 'rb') as f:
            self.env_data = pickle.load(f)

        total_images = len(self.env_data['rgb'])

        if total_images > self.max_images:
            step = max(1, total_images // self.max_images)
            indices = list(range(0, total_images, step))
            self.current_images = [self.env_data['rgb'][i] for i in indices][:self.max_images]
        else:
            self.current_images = self.env_data['rgb']

        self.previous_time_step = None
        self.query_index = 0

    def load_time_step(self, time_step):
        self.query_index = 0
        folder_path = f'env{self.env_number}'
        time_step_file = self.time_steps.get(time_step)

        if not os.path.exists(os.path.join(folder_path, time_step_file)):
            raise FileNotFoundError(f"CSV file {time_step_file} not found.")

        df = pd.read_csv(os.path.join(folder_path, time_step_file))
        total_queries = len(df['query'].dropna())
        self.queries = random.sample(df['query'].dropna().tolist(), min(self.max_queries, total_queries))

        self.current_time_step = time_step
        self.completed_queries = []

        self.load_environment_images()
        self.previous_time_step = time_step

    def clear_memory(self):
        self.env_data = {}
        self.queries = []
        self.completed_queries = []
        self.current_images = []
        self.previous_time_step = None
        gc.collect()


class EnvironmentManager:
    def __init__(self):
        self.environments = {}
        self.shuffled_env_list = []
        self.current_env_index = None
        self.current_env_number = None
        self.current_time_step_index = None
        self.load_environments()

    def load_environments(self):
        env_config = {
            1: {'time_steps': {14: '14.csv', 26: '26.csv', 41: '41.csv'}, 'pkl_file': 'env.pkl'},
            2: {'time_steps': {18: '18.csv', 36: '36.csv', 54: '54.csv'}, 'pkl_file': 'env.pkl'},
            3: {'time_steps': {35: '35.csv', 77: '77.csv', 107: '107.csv', 143: '143.csv'}, 'pkl_file': 'env.pkl'},
            4: {'time_steps': {30: '30.csv', 71: '71.csv', 181: '181.csv'}, 'pkl_file': 'env.pkl'},
            5: {'time_steps': {28: '28.csv', 66: '66.csv', 102: '102.csv'}, 'pkl_file': 'env.pkl'},
            6: {'time_steps': {64: '64.csv', 141: '141.csv', 207: '207.csv'}, 'pkl_file': 'env.pkl'},
            7: {'time_steps': {30: '30.csv', 61: '61.csv', 97: '97.csv'}, 'pkl_file': 'env.pkl'},
            8: {'time_steps': {22: '22.csv', 46: '46.csv', 69: '69.csv'}, 'pkl_file': 'env.pkl'},
            9: {'time_steps': {42: '42.csv', 100: '100.csv', 162: '162.csv'}, 'pkl_file': 'env.pkl'}
        }

        for env_num, config in env_config.items():
            self.environments[env_num] = Environment(env_num, config['time_steps'], config['pkl_file'])

        self.shuffled_env_list = list(self.environments.keys())
        random.shuffle(self.shuffled_env_list)

    def get_next_env_time_step(self):
        if self.current_env_index is None:
            self.current_env_index = 0
            self.current_time_step_index = 0

        current_env_number = self.shuffled_env_list[self.current_env_index]
        env = self.environments[current_env_number]
        time_steps = list(env.time_steps.keys())

        if self.current_time_step_index >= len(time_steps):
            self.current_env_index += 1
            if self.current_env_index >= len(self.shuffled_env_list):
                return {"status": "finished"}
            self.current_time_step_index = 0
            current_env_number = self.shuffled_env_list[self.current_env_index]
            env = self.environments[current_env_number]
            time_steps = list(env.time_steps.keys())

        selected_time_step = time_steps[self.current_time_step_index]
        env.clear_memory()
        env.load_time_step(selected_time_step)
        self.current_time_step_index += 1
        self.current_env_number = current_env_number

        return {
            "status": "success",
            "env_number": env.env_number,
            "time_step": selected_time_step,
            "total_queries": len(env.queries)
        }

env_manager = EnvironmentManager()

def calculate_environment_score(username, env_number):
    directory_path = "/data/mohit/ok-nav/lang-segment-anything/environments_data"
    os.makedirs(directory_path, exist_ok=True)
    file_name = f"{username}.xlsx"
    file_path = os.path.join(directory_path, file_name)
    wb = openpyxl.load_workbook(file_path)
    env_data_sheet = wb['Environment Data']

    total_score = 0
    score_count = 0

    for row in env_data_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == env_number and row[1] == 'Time Step Score':
            total_score += row[8]
            score_count += 1

    env_score = total_score / score_count if score_count > 0 else 0

    headers = [cell.value for cell in env_data_sheet[1]]
    if "Score" in headers:
        score_column_index = headers.index("Score") + 1
        if "Environment Score" not in headers:
            env_data_sheet.insert_cols(score_column_index + 1)
            env_data_sheet.cell(row=1, column=score_column_index + 1, value="Environment Score")

        for row_idx, row in enumerate(env_data_sheet.iter_rows(min_row=2, max_row=env_data_sheet.max_row), start=2):
            if row[0] == env_number:
                env_data_sheet.cell(row=row_idx, column=score_column_index + 1, value=env_score)

    wb.save(file_path)
    return env_score

def calculate_final_score(username):
    file_name = f"{username}.xlsx"
    file_path = os.path.join(os.getcwd(), file_name)
    wb = openpyxl.load_workbook(file_path)
    env_data_sheet = wb['Environment Data']

    total_score = 0
    env_count = 0

    for row in env_data_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == 'Environment Score':
            total_score += row[8]
            env_count += 1

    final_score = total_score / env_count if env_count > 0 else 0
    env_data_sheet.append(['Final Score', '', '', '', '', '', '', '', final_score])
    wb.save(file_path)
    return final_score


@app.route('/next_env_time_step', methods=['GET'])
def next_env_time_step():
    try:
        result = env_manager.get_next_env_time_step()
        if result['status'] == 'finished':
            return jsonify({"status": "finished"})
        return jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)})


@app.route('/get_next_query')
def get_next_query():
    if env_manager.current_env_number is None:
        return jsonify({"status": "error", "message": "No environment is currently loaded"}), 400

    env = env_manager.environments.get(env_manager.current_env_number)
    if env is None:
        return jsonify({"status": "error", "message": "Invalid environment"}), 400

    query_index = env.query_index
    if query_index < len(env.queries):
        query = env.queries[query_index]
        env.query_index += 1
        username = request.args.get('username')
        if username:
            save_user_progress(username, env.env_number, env.current_time_step, env.query_index)

        return jsonify({
            "query": query,
            "finished": False,
            "completed": query_index,
            "remaining": len(env.queries) - query_index
        })
    else:
        return jsonify({"query": "", "finished": True})


@app.route('/skip_query', methods=['POST'])
def skip_query():
    try:
        env = env_manager.environments.get(env_manager.current_env_number)
        if not env:
            return jsonify({"status": "error", "message": "Environment not found."}), 404

        if not (0 < env.query_index <= len(env.queries)):
            return jsonify({"status": "error", "message": "Invalid query index."}), 400

        query = env.queries[env.query_index - 1]
        data = request.get_json()
        username = data.get('username')
        if not username:
            return jsonify({"status": "error", "message": "Username is missing."}), 400

        folder_path = f'env{env.env_number}'
        voxel_coords = [None, None, None]
        success = False

        csv_file = env.time_steps.get(env.current_time_step)
        df = pd.read_csv(os.path.join(folder_path, csv_file))
        if query in df['query'].values:
            row = df[df['query'] == query].iloc[0]
            voxel_coords = row[1:4].values.astype(float).tolist()
            if all(pd.isna(voxel_coords)):
                success = True
        else:
            raise ValueError(f"Query '{query}' not found in the current time step.")

        voxel_coords = [coord if not pd.isna(coord) else None for coord in voxel_coords]

        environment_data = [{
            'environment': env.env_number,
            'time_step': env.current_time_step,
            'query': query,
            'response': 'Not Found',
            'manual_coordinates': None,
            'voxel_map_coordinates': voxel_coords,
            'elapsed_time': data.get('elapsed_time'),
            'result': 'Success' if success else 'Skipped',
            'score': 1 if success else 0
        }]

        save_environment_data_to_excel(username, environment_data, env)
        save_user_progress(username, env.env_number, env.current_time_step, env.query_index)

        return jsonify({"status": "success", "message": f"Query '{query}' skipped.", "score_awarded": 1 if success else 0})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


def save_environment_data_to_excel(username, environment_data, env):
    if not username:
        raise ValueError("Username is missing.")
    if not isinstance(environment_data, list) or not environment_data:
        raise ValueError("Environment data must be a non-empty list.")

    file_name = f"{username}.xlsx"
    file_path = os.path.join(os.getcwd(), file_name)

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()

    if 'Environment Data' not in wb.sheetnames:
        wb.create_sheet('Environment Data')

    env_data_sheet = wb['Environment Data']

    if env_data_sheet.max_row == 1:
        env_data_sheet.append([
            'Environment', 'Time Step', 'Query', 'Response',
            'Manual Coordinates', 'Voxel Map Coordinates',
            'Elapsed Time', 'Result', 'Score'
        ])

    for data in environment_data:
        env_data_sheet.append([
            data.get('environment', ''),
            data.get('time_step', ''),
            data.get('query', ''),
            data.get('response', 'Skipped'),
            str(data.get('manual_coordinates')) if data.get('manual_coordinates') else None,
            str(data.get('voxel_map_coordinates')) if data.get('voxel_map_coordinates') else None,
            data.get('elapsed_time', None),
            data.get('result', 'Skipped'),
            data.get('score', 0)
        ])

    if not hasattr(env, 'time_step_scores'):
        env.time_step_scores = []

    if env.query_index >= len(env.queries):
        total_ones = 0
        total_zeros = 0

        for row in env_data_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == data['time_step']:
                score_value = row[8]
                if score_value == 1:
                    total_ones += 1
                elif score_value == 0:
                    total_zeros += 1

        total_queries = total_ones + total_zeros
        time_step_score = (total_ones / total_queries) * 100 if total_queries > 0 else 0
        env.time_step_scores.append(time_step_score)

        env_data_sheet.append(['Time Step Score', '', '', '', '', '', '', '', time_step_score])

    total_env_ones = 0
    total_env_zeros = 0

    for row in env_data_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == env.env_number:
            score_value = row[8]
            if score_value == 1:
                total_env_ones += 1
            elif score_value == 0:
                total_env_zeros += 1

    total_env_queries = total_env_ones + total_env_zeros
    environment_score = (total_env_ones / total_env_queries) * 100 if total_env_queries > 0 else 0

    env_data_sheet.append(['Environment Score', '', '', '', '', '', '', '', environment_score])
    wb.save(file_path)


def save_user_progress(username, env_number, time_step, query_index):
    progress_file = f"{username}_progress.pkl"
    progress_data = {
        'env_number': env_number,
        'time_step': time_step,
        'query_index': query_index
    }
    with open(progress_file, 'wb') as f:
        pickle.dump(progress_data, f)


@app.route('/save_pause_reason', methods=['POST'])
def save_pause_reason():
    try:
        data = request.json
        username = data['username']
        env_number = data['env_number']
        time_step = data['time_step']
        query = data['query']
        pause_reason = data['pause_reason']

        file_name = f"{username}.xlsx"
        file_path = os.path.join(os.getcwd(), file_name)

        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()

        if 'Environment Data' not in wb.sheetnames:
            wb.create_sheet('Environment Data')

        env_data_sheet = wb['Environment Data']

        headers = [cell.value for cell in env_data_sheet[1]]

        if "Pause Reason" not in headers:
            pause_reason_column = len(headers) + 1
            env_data_sheet.cell(row=1, column=pause_reason_column, value="Pause Reason")
        else:
            pause_reason_column = headers.index("Pause Reason") + 1

        query_found = False
        for row in env_data_sheet.iter_rows(min_row=2, max_col=pause_reason_column):
            if row[0].value == env_number and row[1].value == time_step and row[2].value == query:
                env_data_sheet.cell(row=row[0].row, column=pause_reason_column, value=pause_reason)
                query_found = True
                break

        if not query_found:
            env_data_sheet.append([env_number, time_step, query, '', '', '', '', '', '', pause_reason])

        wb.save(file_path)

        return jsonify({"status": "success", "message": "Pause reason saved."})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


def get_xyz(depth, pose, intrinsics):
    if not isinstance(depth, torch.Tensor):
        depth = torch.from_numpy(depth)
    if not isinstance(pose, torch.Tensor):
        pose = torch.from_numpy(pose)
    if not isinstance(intrinsics, torch.Tensor):
        intrinsics = torch.from_numpy(intrinsics)

    while depth.ndim < 4:
        depth = depth.unsqueeze(0)
    while pose.ndim < 3:
        pose = pose.unsqueeze(0)
    while intrinsics.ndim < 3:
        intrinsics = intrinsics.unsqueeze(0)

    (bsz, _, height, width), device, dtype = depth.shape, depth.device, intrinsics.dtype

    xs, ys = torch.meshgrid(
        torch.arange(0, width, device=device, dtype=dtype),
        torch.arange(0, height, device=device, dtype=dtype),
        indexing="xy",
    )
    xy = torch.stack([xs, ys], dim=-1).flatten(0, 1).unsqueeze(0).repeat_interleave(bsz, 0)
    xyz = torch.cat((xy, torch.ones_like(xy[..., :1])), dim=-1)
    xyz = xyz @ get_inv_intrinsics(intrinsics).transpose(-1, -2)
    xyz = xyz * depth.flatten(1).unsqueeze(-1)
    xyz = (xyz[..., None, :] * pose[..., None, :3, :3]).sum(dim=-1) + pose[..., None, :3, 3]
    xyz = xyz.unflatten(1, (height, width))

    return xyz


def get_inv_intrinsics(intrinsics):
    return torch.inverse(intrinsics)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/save_user_info', methods=['POST'])
def save_user_info():
    try:
        data = request.json
        username = data['username']
        email = data['email']

        if ' ' in username:
            return jsonify({"status": "error", "message": "Username cannot contain spaces."})

        file_name = f"{username}.xlsx"
        file_path = os.path.join(os.getcwd(), file_name)

        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()

        if 'User Info' not in wb.sheetnames:
            wb.create_sheet('User Info')

        user_info_sheet = wb['User Info']
        user_info_sheet.delete_rows(1, user_info_sheet.max_row)

        user_info_sheet.append(['Field', 'Information'])
        user_info_sheet.append(['Username', username])
        user_info_sheet.append(['Email', email])

        wb.save(file_path)

        environment_data = []
        if environment_data:
            save_environment_data_to_excel(username, environment_data, env)

        return jsonify({"status": "success", "file": file_name})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


@app.route('/start_timer', methods=['POST'])
def start_timer():
    global start_time
    start_time = time.time()
    return jsonify({"status": "timer_started", "start_time": start_time})


@app.route('/select_env', methods=['POST'])
def select_env():
    data = request.json
    env_number = data['env_number']
    time_step = data.get('time_step')

    env = env_manager.environments.get(int(env_number))
    if not env:
        return jsonify({"status": "error", "message": "Environment not found."})

    if time_step is None:
        time_steps_with_queries = {}
        folder_path = f'env{env_number}'

        for ts, csv_file in env.time_steps.items():
            df = pd.read_csv(os.path.join(folder_path, csv_file))
            queries = df['query'].dropna().tolist()
            time_steps_with_queries[ts] = queries

        return jsonify({"status": "time_step_selection", "time_steps": time_steps_with_queries})

    try:
        env.load_time_step(time_step)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


@app.route('/get_images', methods=['POST'])
def get_images():
    try:
        data = request.json
        page_number = data.get('page_number', 0)
        page_size = data.get('page_size', 24)

        env = env_manager.environments.get(env_manager.current_env_number)
        if not env:
            return jsonify({"status": "error", "message": "Environment not found."})

        current_images = env.current_images
        if not current_images:
            return jsonify({"status": "error", "message": "No images available for the current environment."})

        start_index = page_number * page_size
        end_index = min(start_index + page_size, len(current_images))

        images = []
        for i in range(start_index, end_index):
            image_tensor = current_images[i]

            if isinstance(image_tensor, torch.Tensor):
                image_array = image_tensor.cpu().numpy()
            else:
                image_array = image_tensor

            pil_image = Image.fromarray(image_array.astype('uint8'))
            buffer = io.BytesIO()
            pil_image.save(buffer, format="JPEG")
            encoded_image = base64.b64encode(buffer.getvalue()).decode('utf-8')

            images.append({
                "image": encoded_image,
                "index": i
            })

        return jsonify({"images": images})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


@app.route('/save_points', methods=['POST'])
def save_points():
    try:
        data = request.json
        query = data['query']
        points = data['points']
        env_number = data['env_number']
        username = data.get('username')

        if not username:
            raise ValueError("User information is missing")

        env = env_manager.environments.get(env_number)
        if not env:
            raise ValueError(f"Environment {env_number} not found in the configuration")

        time_step = data.get('time_step', env.current_time_step)
        current_env_data = env.env_data
        completed_queries = env.completed_queries

        folder_path = f'env{env_number}'
        xyz_coordinates = []
        tolerance_satisfied = False
        voxel_coords = [None, None, None]

        for point in points:
            image_index = point['image_index']
            depth = current_env_data['depth'][image_index]
            pose = current_env_data['camera_poses'][image_index]
            intrinsics = current_env_data['camera_K'][image_index]

            xyz = get_xyz(depth, pose, intrinsics)
            if xyz.ndim == 4 and xyz.shape[0] == 1:
                xyz = xyz.squeeze(0)
            scaled_x = int(point['x'] * depth.shape[1])
            scaled_y = int(point['y'] * depth.shape[0])
            xyz_point = xyz[scaled_y, scaled_x].tolist()
            rounded_xyz = [round(coord, 2) for coord in xyz_point]
            xyz_coordinates.append({"2d": point, "3d": rounded_xyz})

            csv_file = env.time_steps.get(env.current_time_step)
            df = pd.read_csv(os.path.join(folder_path, csv_file))
            if query in df['query'].values:
                row = df[df['query'] == query].iloc[0]
                voxel_coords = row[1:4].values.astype(float).tolist()
                tolerance = float(row.iloc[4]) if len(row) > 4 else 0
                distance = sum([(rounded_xyz[i] - voxel_coords[i]) ** 2 for i in range(3)]) ** 0.5
                if distance <= tolerance:
                    tolerance_satisfied = True
            else:
                raise ValueError(f"Query '{query}' not found in the current time step.")

        voxel_coords = [coord if not pd.isna(coord) else None for coord in voxel_coords]

        selected_object = f"Selected: {xyz_coordinates[0]['3d']}" if xyz_coordinates else 'None'
        environment_data = [{
            'environment': env_number,
            'time_step': time_step,
            'query': query,
            'response': selected_object,
            'manual_coordinates': xyz_coordinates[0]['3d'] if xyz_coordinates else None,
            'voxel_map_coordinates': voxel_coords,
            'elapsed_time': data.get('elapsed_time', None),
            'result': 'Success' if tolerance_satisfied else 'Failure',
            'score': 1 if tolerance_satisfied else 0
        }]

        save_environment_data_to_excel(username, environment_data, env)

        if tolerance_satisfied:
            completed_queries.append(query)
            return jsonify({"status": "success", "xyz_coordinates": xyz_coordinates, "tolerance_satisfied": True})
        else:
            return jsonify({"status": "success", "xyz_coordinates": xyz_coordinates, "tolerance_satisfied": False})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)})


@app.route('/get_query_coordinates', methods=['POST'])
def get_query_coordinates():
    try:
        data = request.json
        query = data['query']
        env_number = data['env_number']
        time_step = data.get('time_step', None)

        env_info = env_manager.environments.get(env_number)
        if env_info is None:
            raise ValueError(f"Environment {env_number} not found in configuration.")

        folder_path = f'env{env_number}'

        if time_step is None:
            time_step = env_manager.current_time_step

        csv_file = env_info.time_steps.get(time_step)
        if not csv_file:
            raise FileNotFoundError(f"CSV file not found for time step {time_step} in environment {env_number}")

        csv_path = os.path.join(folder_path, csv_file)

        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"CSV file {csv_file} not found at {csv_path}")

        df = pd.read_csv(csv_path)

        if query not in df['query'].values:
            raise ValueError(f"Query '{query}' not found in time step {time_step} for environment {env_number}")

        row = df[df['query'] == query].iloc[0]
        absolute_coordinates = row.iloc[1:].tolist()

        rounded_coordinates = [round(float(coord), 6) for coord in absolute_coordinates[:3]] + [float(absolute_coordinates[3])]

        return jsonify({"status": "success", "xyz_coordinates": rounded_coordinates})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5007)
